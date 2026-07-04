"""Excel → markdown 表格 converter。

每個 sheet 一段 `## <sheet名>` ＋ md 表格；合併儲存格取左上值、其餘補空；
超過 MAX_ROWS × MAX_COLS 截斷並於 md 尾註記。
"""

from pathlib import Path

from openpyxl import load_workbook

from app.ingest.dispatcher import IngestError, IngestResult
from app.store.project import Project

MAX_ROWS = 200
MAX_COLS = 30

_TRUNCATED_NOTE = "（表格過大，已截斷：僅保留前 {rows} 列 × {cols} 欄）"


def _format_cell(value) -> str:
    """儲存格值轉字串：None 補空、int 保持整數、避免 42 變 42.0。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("|", "\\|").replace("\n", " ")


def _sheet_to_markdown(ws) -> tuple[str, list[str]]:
    """單一 worksheet 轉 md 表格，回傳 (markdown, warnings)。"""
    warnings: list[str] = []
    n_rows = ws.max_row or 0
    n_cols = ws.max_column or 0

    truncated = n_rows > MAX_ROWS or n_cols > MAX_COLS
    n_rows = min(n_rows, MAX_ROWS)
    n_cols = min(n_cols, MAX_COLS)

    if n_rows == 0 or n_cols == 0:
        return "（空白工作表）", warnings

    lines: list[str] = []
    for r, row in enumerate(
        ws.iter_rows(min_row=1, max_row=n_rows, min_col=1, max_col=n_cols)
    ):
        # MergedCell 的 value 為 None（左上主儲存格才有值），自然補空
        cells = [_format_cell(cell.value) for cell in row]
        lines.append("| " + " | ".join(cells) + " |")
        if r == 0:
            lines.append("|" + " --- |" * n_cols)

    md = "\n".join(lines)
    if truncated:
        note = _TRUNCATED_NOTE.format(rows=MAX_ROWS, cols=MAX_COLS)
        md += f"\n\n{note}"
        warnings.append(f"{ws.title}：{note}")
    return md, warnings


def convert_excel(src: Path, project: Project) -> IngestResult:
    """xlsx/xlsm 轉 markdown，輸出到 project/md/<原檔名>.md。"""
    try:
        wb = load_workbook(src, data_only=True, read_only=False)
    except Exception as exc:
        raise IngestError(f"無法開啟 Excel 檔案（檔案可能已損壞）：{src.name}") from exc

    warnings: list[str] = []
    sections: list[str] = []
    try:
        for ws in wb.worksheets:
            table_md, ws_warnings = _sheet_to_markdown(ws)
            sections.append(f"## {ws.title}\n\n{table_md}")
            warnings.extend(ws_warnings)
    finally:
        wb.close()

    dest = project.path / "md" / f"{src.name}.md"
    try:
        dest.write_text("\n\n".join(sections) + "\n", encoding="utf-8")
    except OSError as exc:
        raise IngestError(f"寫入轉換結果失敗：{dest.name}") from exc

    return IngestResult(
        src_name=src.name,
        output_type="markdown",
        output_path=dest,
        warnings=warnings,
    )
