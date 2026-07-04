"""Ingest dispatcher 與 md/excel converter 的基本測試。"""

import pytest
from openpyxl import Workbook

from app.ingest import (
    IngestError,
    UnsupportedFormatError,
    ingest_file,
)
from app.ingest.excel_converter import MAX_COLS, MAX_ROWS
from app.store.project import create_project


@pytest.fixture
def project(tmp_path):
    return create_project(tmp_path / "projects", "ingest 測試")


@pytest.fixture
def sample_xlsx(tmp_path):
    """兩欄三列、含合併儲存格與數字的 xlsx fixture。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "報表"
    ws["A1"] = "項目"
    ws["B1"] = "數量"
    ws["A2"] = "蘋果"
    ws["B2"] = 42
    ws["A3"] = "合併列"
    ws.merge_cells("A3:B3")
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return path


# ---------- dispatcher ----------


def test_unsupported_extension_raises_with_supported_list(tmp_path, project):
    src = tmp_path / "data.csv"
    src.write_text("a,b\n1,2\n", encoding="utf-8")
    with pytest.raises(UnsupportedFormatError) as excinfo:
        ingest_file(src, project)
    assert ".md" in str(excinfo.value)
    assert ".xlsx" in str(excinfo.value)


def test_xls_is_unsupported(tmp_path, project):
    src = tmp_path / "legacy.xls"
    src.write_bytes(b"\x00\x01")
    with pytest.raises(UnsupportedFormatError):
        ingest_file(src, project)


def test_unsupported_error_is_ingest_error():
    assert issubclass(UnsupportedFormatError, IngestError)


def test_docx_planned_but_not_yet_supported(tmp_path, project):
    src = tmp_path / "doc.docx"
    src.write_bytes(b"PK")
    with pytest.raises(IngestError, match="尚未支援") as excinfo:
        ingest_file(src, project)
    # 已規劃格式不應被歸類為「不支援的副檔名」
    assert not isinstance(excinfo.value, UnsupportedFormatError)


def test_pdf_planned_but_not_yet_supported(tmp_path, project):
    src = tmp_path / "paper.pdf"
    src.write_bytes(b"%PDF")
    with pytest.raises(IngestError, match="尚未支援"):
        ingest_file(src, project)


# ---------- markdown / txt ----------


def test_md_ingest_copies_content_verbatim(tmp_path, project):
    src = tmp_path / "notes.md"
    content = "# 標題\n\n- 重點一\n- 重點二\n"
    src.write_text(content, encoding="utf-8")

    result = ingest_file(src, project)

    assert result.src_name == "notes.md"
    assert result.output_type == "markdown"
    assert result.output_path == project.path / "md" / "notes.md.md"
    assert result.output_path.read_text(encoding="utf-8") == content
    assert result.warnings == []


def test_txt_treated_as_markdown(tmp_path, project):
    src = tmp_path / "memo.txt"
    src.write_text("純文字內容\n", encoding="utf-8")

    result = ingest_file(src, project)

    assert result.output_type == "markdown"
    assert result.output_path == project.path / "md" / "memo.txt.md"
    assert result.output_path.read_text(encoding="utf-8") == "純文字內容\n"


def test_md_ingest_rejects_non_utf8(tmp_path, project):
    src = tmp_path / "legacy.md"
    src.write_bytes("Big5 編碼內容".encode("big5"))
    with pytest.raises(IngestError) as excinfo:
        ingest_file(src, project)
    msg = str(excinfo.value)
    assert "legacy.md" in msg
    assert "UTF-8" in msg
    assert "Traceback" not in msg


def test_ingest_result_extra_assets_defaults_to_empty_tuple(tmp_path, project):
    src = tmp_path / "notes.md"
    src.write_text("內容\n", encoding="utf-8")
    result = ingest_file(src, project)
    assert result.extra_assets == ()


# ---------- excel ----------


def test_xlsx_converts_to_markdown_table(project, sample_xlsx):
    result = ingest_file(sample_xlsx, project)

    assert result.output_type == "markdown"
    assert result.output_path == project.path / "md" / "sample.xlsx.md"
    md = result.output_path.read_text(encoding="utf-8")

    assert "## 報表" in md
    assert "| 項目 | 數量 |" in md
    # 數字保留原值：int 不能變 42.0
    assert "| 蘋果 | 42 |" in md
    assert "42.0" not in md
    # 合併儲存格：左上取值、其餘補空
    assert "| 合併列 |  |" in md
    assert result.warnings == []


def test_xlsx_truncates_over_limit(tmp_path, project):
    wb = Workbook()
    ws = wb.active
    ws.title = "big"
    for r in range(MAX_ROWS + 5):
        for c in range(MAX_COLS + 3):
            ws.cell(row=r + 1, column=c + 1, value=r * 1000 + c)
    path = tmp_path / "big.xlsx"
    wb.save(path)

    result = ingest_file(path, project)
    md = result.output_path.read_text(encoding="utf-8")

    assert "已截斷" in md
    assert any("已截斷" in w for w in result.warnings)
    # 行數不超過上限（header 分隔線與標題、註記除外）
    table_rows = [line for line in md.splitlines() if line.startswith("|")]
    assert len(table_rows) <= MAX_ROWS + 1  # +1 為分隔線


def test_format_cell_strips_carriage_returns():
    from app.ingest.excel_converter import _format_cell

    assert _format_cell("第一行\r\n第二行") == "第一行  第二行"
    assert _format_cell("尾端\r") == "尾端 "


def test_corrupt_xlsx_raises_friendly_ingest_error(tmp_path, project):
    src = tmp_path / "broken.xlsx"
    src.write_bytes(b"this is not a zip archive")
    with pytest.raises(IngestError) as excinfo:
        ingest_file(src, project)
    msg = str(excinfo.value)
    assert "broken.xlsx" in msg
    assert "Traceback" not in msg


# ---------- images ----------


def test_image_copied_to_assets(tmp_path, project):
    src = tmp_path / "chart.png"
    src.write_bytes(b"\x89PNG\r\n\x1a\n fake")

    result = ingest_file(src, project)

    assert result.output_type == "asset"
    assert result.output_path == project.path / "assets" / "chart.png"
    assert result.output_path.read_bytes() == src.read_bytes()


def test_duplicate_image_name_gets_suffix(tmp_path, project):
    src1 = tmp_path / "chart.png"
    src1.write_bytes(b"first")
    r1 = ingest_file(src1, project)

    src2 = tmp_path / "sub"
    src2.mkdir()
    dup = src2 / "chart.png"
    dup.write_bytes(b"second")
    r2 = ingest_file(dup, project)

    assert r1.output_path != r2.output_path
    assert r1.output_path.read_bytes() == b"first"  # 不覆蓋
    assert r2.output_path.read_bytes() == b"second"
    assert r2.output_path.parent == project.path / "assets"
    assert r2.output_path.name == "chart-1.png"
